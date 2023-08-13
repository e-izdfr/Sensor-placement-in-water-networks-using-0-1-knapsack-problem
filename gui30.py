import PySimpleGUI as sg  # Python Package for making GUIs: It is a combination of the former packages (for instance tkinter) and easier to work with.
import pandas as pd  # Python package for dealing with tabular data(Data Manipulation)
import numpy as np  # Python package for dealing with arrays(Data Manipulation)
import openpyxl  # Python library for reading and writing to excel files
import pulp as pl  # Python package for modeling linear programming problems
import os  # Python module for operating-system-related matters

def isfloat(string):  # This function receives a string 'string' and returns True if it is a float and returns False if it is not.
    try: 
        float(string)  # Checking whether the string is a float or not
    except ValueError:  # Checking whether 'ValueError' is received or not
        return False
    return True

def update(element, text):
    element.TooltipObject.text = text
os.chdir(os.path.realpath(__file__).removesuffix('\\gui30.py'))
ICON = 'LogoProDesign2.ico'

######################################################################################################################################################

'''
In this segment, a GUI is made and different features of its layout(i.e widgets) are adjusted according to circumstances of the problem.
'''
#sg.theme_previewer()
#sg.theme('Dark')
window = sg.Window(  # Making a GUI window(object) named 'window' with the '.Window' command and specifying its attributes(widgets)
    title='                                                                                                         Sensor placement in large industrial networks',  
    # Giving the window a title with the 'title' variable
    layout=[  # Specifying the different features of the layout(widgets) with the 'layout' variable
        [sg.Text(text='Enter the number of rows of your excel file:'),  # Displaying a text(object) with the '.Text' command: The user 
         # should enter the number of rows of his excel file without counting the header row.
         sg.InputText(size =(15, 1), key='IN1')],  # Receiving a string from the user with the '.InputText' command and giving it a key named 'IN1': Note that all the inputs from the user 
         # are automatically stored in a dictionary so we need to assign a key to the input text.
        [sg.Text('Enter your budget:'),  # Receiving the current total budget as a string
        sg.InputText(size =(15, 1), key='IN3')],  # Specifying a key 'IN3' for the input string
        [sg.Text("Click to see the sample excel file. Don't change the path and the name of the file."),  # Some
         # information about the buttons we're about to make
         sg.Button('Click'),  # Adding a button(object) named 'Click' with the .Button command whose clicking makes the sample excel file appear
         sg.Button('Save', visible=False)],  # Adding a button named 'Save' whose clicking saves the sample excel file and making it invisible with 'visible' variable
        [sg.Input(key='_FILEBROWSE_', enable_events=True, visible=False)],  # Adding this event-making command since clicking the 'FileBrowse' button is not considered an
        # event, i.e clicking this button as an input becomes an event.
        [sg.Text('Choose your excel file:'),  # Displaying another text
        sg.FileBrowse(key='IN2', file_types=(('EXCEL FILES', '*.xlsx'),), target='_FILEBROWSE_', tooltip='Choose your excel file:')],  # Receiving a file path from the user with the '.FileBrowse' command:
         # Notice that a key 'IN2' is assigned and the legal input file types have been restricted to excel files with the 'file_types' variable. The 'target' variable assures
         # that clicking this button is considered an event.
        [sg.Button('OK', disabled=True)],  # Making an 'OK' button whose clicking starts the validation and optimization process and making it disabled at first with
        # 'disabled' variable
        [sg.ProgressBar(max_value=100, orientation='horizontal', size=(10, 10), visible = False, bar_color=('green', 'white'), style='vista', key ='k')]],  # Making a progress bar with '.ProgreesBar' method and
        # specifying its size with 'size' variable, its orientation with 'orientation' variable, the capacity of the bar with 'max_value' variable, its visibility status with
        # 'visible' variable and giving it a key 'k'
    margins=(200, 200), icon=ICON, element_justification='left', finalize=True)
window['Click'].set_cursor('hand2')  # Changing the cursor of the 'Click' button to a hand
window['IN2'].set_cursor('hand2')  # Changing the cursor of the 'IN2' input text to a hand
event, values = window.read()  # '.read' method executes the input and event assignments. Events are things like clicking a button,choosing an element from a list,etc.
if event is not None:
    update(window['IN2'], values['IN2'])
##############################################################################################################################################################################

'''
In this part, the sample excel file is made and the validation tests are implemented on it.
'''

key = 1  # Initializing the key for understanding whether there is any problem with the different columns or not
key2 = 0  # Initializing the key for understanding whether the user has both filled tha sample excel sheet and browsed an excel file, or not
key3 = 1  # Initializing the key for understanding whether the progress bar should be made visible or not
key5 = 1  # Initializing the key for understanding whether the first row(i.e importance coefficients) of the input excell file is empty or not
key6 = 1  # Initializing the key for understanding whether any empty cell is found or not

while True:
    while event == 'Click':  # Checking whether the user has clicked the button 'Click' or not
        workbook = openpyxl.Workbook()  # Making a workbook(object) named 'workbook' with '.Workbook' command
        worksheet = workbook.active  # Making a worksheet named 'worksheet' with '.active' method, i.e setting 'workbook' as the active worksheet
        c11 = worksheet.cell(row=1, column=1)  # Creating the 'c11' cell(object) of the sheet with '.cell' method
        c11.value = 'لوله ھای بعد از مخزن یا پمپ'  # Assigning value to 'c11' with '.value' variable
        c12 = worksheet.cell(row=1, column=2)  # Creating the 'c12' cell of the sheet
        c12.value = 'لوله ھای قبل یا بعد شیر فشارشکن'  # Assigning value to 'c12'
        c13 = worksheet.cell(row=1, column=3)  # Creating the 'c13' cell of the sheet
        c13.value = 'مشترکین پرمصرف'  # Assigning value to 'c13'
        c14 = worksheet.cell(row=1, column=4)  # Creating the 'c14' cell of the sheet
        c14.value = 'نقاط حساس شبکه'  # Assigning value to 'c14'
        c15 = worksheet.cell(row=1, column=5)  # Creating the 'c15' cell of the sheet
        c15.value = 'محل انشعابات روستایی'  # Assigning value to 'c15'
        c16 = worksheet.cell(row=1, column=6)  # Creating the 'c16' cell of the sheet
        c16.value = 'لوله ھای مجاور با شیر کنترل دبی'  # Assigning value to 'c16'
        c17 = worksheet.cell(row=1, column=7)  # Creating the 'c17' cell of the sheet
        c17.value = 'هزینه نصب'  # Assigning value to 'c17'
        c18 = worksheet.cell(row=1, column=8)  # Creating the 'c18' cell of the sheet
        c18.value = 'نمايه'  # Assigning value to 'c18'
        try:
            workbook.save('sample.xlsx')  # Saving the workbook as an excel file 'sample.xlsx' with '.save' method
            os.startfile('sample.xlsx')  # Opening the file 'sample.xlsx' with '.startfile' command
        
        except:
            sg.popup('Another excel file is open right now. Close it.', title='', background_color='blue', icon=ICON)
        
        event = 'Save'  # Saving the excell file(from our view point) since the user has clicked the 'Click' button
        if event == 'Save':  # Checking whether the file has been saved or not(This is always true. Note that there is an invisible 'save' button. Check the layout.)
            key2 = 1  # Changing the value of 'key2' since the user has saved the filled excel file
            if key2 == 1:
                window['OK'].Update(disabled=False)  # Finding the 'OK' button with '.FindElement' method and updating the 'disabled' attribute with '.Update'
                # method: Note that the 'FileBrowse' button has been made visible since the user has filled a file.
                window['OK'].set_cursor('hand2')  # Changing the cursor of the 'OK' button to a hand
                #sg.ProgressBar.update(window['k'], visible=True, current_count=0)  # Updating the attributes of the bar with '.update' method: Note that
                # 'current_count' is the percentage to which the bar has filled and it has been made visible since the user has filled a file.
            event, values = window.read()  # Receiving input again
            if values != None:
                update(window['IN2'], values['IN2'])     

            while event == 'OK':  # Checking whether the user has pressed the 'OK' button or not
                if event == 'OK' and values == None:  # Checking whether the user has closed the gui window or not
                    break  # Getting out of the event loop
                workbook = openpyxl.load_workbook('sample.xlsx')  # Loading the workbook named 'sample.xlsx' with '.load_workbook' method onto 'workbook'
                worksheet = workbook.active  # Making a worksheet named 'worksheet'
                worksheet.cell(row=1, column=1).value = 'after_pump'  # Changing the name of the first column header
                worksheet.cell(row=1, column=2).value = 'pressure_relief'  # Changing the name of the second column header
                worksheet.cell(row=1, column=3).value = 'high consumers'  # Changing the name of the third column header
                worksheet.cell(row=1, column=4).value = 'critical points'   # Changing the name of the fourth column header
                worksheet.cell(row=1, column=5).value = 'village bifurcations'  # Changing the name of the fifth column header
                worksheet.cell(row=1, column=6).value = 'control valve'  # Changing the name of the sixth column header
                worksheet.cell(row=1, column=7).value = 'installation cost'  # Changing the name of the seventh column header
                worksheet.cell(row=1, column=8).value = 'index'  # Changing the name of the eighth column header
                workbook.save('!!!.xlsx')  # Saving the workbook as an excel file
                key6 = 1  # Updating the value of 'key6'
                key = 1  # Initializing 'key' again just in case
                table = pd.read_excel('!!!.xlsx', header=0)  # Reading the input excel file with '.read_excel' command and setting its first row as the header(labels row) 
                # with 'header' variable : Note that 'table' is a dataframe(object)
                table.columns = ['after_pump', 'pressure_relief', 'high consumers', 'critical points', 'village bifurcations', 'control valve', 'installation cost', 'index']
                # Updating the header with '.columns' attribute
                length = len(table)  # Finding the index of the last row at which there is an entry
                if length == 0:  # Checking whether the number of rows in the excell file is zero or not
                    window['OK'].Update(disabled=True)  # Making the 'OK' button invisible since the user hasn't saved the excell file
                    window['OK'].set_cursor('arrow')  # Changing the cursor of the 'OK' button to a hand
                    sg.ProgressBar.update(window['k'], visible=False, current_count=0)  # Making the progress bar invisible since the user hasn't saved the
                    # excell file
                    window2 = sg.popup("You didn't save the filled sample excel file.", title='', background_color='blue', icon=ICON)  # Displaying the relevant message
                    event , values = window.read()  # Receiving input again
                    if values != None:
                        update(window['IN2'], values['IN2'])
                    if event == '_FILEBROWSE_':  # Checking whether the user has clicked the 'FileBrowse' button or not
                        window['OK'].Update(disabled=False)  # Making the 'OK' button visible since the user has browsed a file
                        window['OK'].set_cursor('hand2')  # Changing the cursor of the 'OK' button to a hand
                        event = '_FILEBROWSE_'  # Updating the value of event since we want exit the current loop to go to the browsed file loop
                    continue  # Going back to the validation cycle

                step = 0  # Initializing the key for understanding how many empty rows(other than the first row) there are in the excell file
                J = []  # Initializing the list containing the indices of the empty rows of the file
                for i in range(1, length):
                    if table.iloc[[i], :].isnull().values.all():  # Checking whether the (i + 1)th row is empty or not: '.iloc' function is used to access the different values
                        # of different cells of the dataframe and '.isnull()' method returns a boolean dataframe indicating whether different cells are empty or not and
                        #  '.value' turns this dataframe into an array and '.all()' checks whether all the elements of this array are 'True' or not
                        step += 1  # Incrementing by 1
                        J.append(i)  # Adding this index to 'J'

                if (not values['IN1'].isnumeric() and values['IN1'] != '') or values['IN1'] == '0':  # Checking whether the input is a natural number or not
                    window2 = sg.popup('The number of rows is invalid.', title='', background_color='blue', icon=ICON)  # Displaying the relevant message
                    key = 0  # Updating the value of 'key'
                    event, values = window.read()  # Receiving input again
                    if values != None:
                        update(window['IN2'], values['IN2'])    
                    continue  # Going back to the validation cycle

                elif values['IN1'] != '' and values['IN1'] is not None:  # Making sure that the user has entered the number of rows and hasn't closed the window
                    if (length - step - 1)  != int(values['IN1']):  # Checking whether the user has counted the number of rows correctly or not
                        window2 = sg.popup('You made a mistake in counting the number of rows.', title='', background_color='blue', icon=ICON)  # Displaying the relevant message
                        window2 = sg.popup('Its number is', length - step - 1, '.', title='', background_color='blue', icon=ICON)
                        key = 0  # Updating the value of 'key'
                        event, values = window.read()  # Receiving input again
                        if values != None:
                            update(window['IN2'], values['IN2'])     
                        continue  # Going back to the validation cycle

                else:
                    window2 = sg.popup("You didn't enter the number of rows.", title='', background_color='blue', icon=ICON)  # Displaying the relevant message
                    key = 0  # Updating the value of 'key'
                    event, values = window.read()  # Receiving input again
                    if values != None:
                        update(window['IN2'], values['IN2']) 
                    continue  # Going back to the validation cycle

                if (not isfloat(values['IN3']) and values['IN3'] != '') or (isfloat(values['IN3']) and float(values['IN3']) < 0):  # Checking whether the input is a nonnegative
                    # real number or not
                    window2 = sg.popup('Your budget is invalid.', title='', background_color='blue', icon=ICON)  # Displaying the relevant message
                    key = 0  # Updating the value of 'key'
                    event, values = window.read()  # Receiving input again
                    if values != None:
                        update(window['IN2'], values['IN2']) 
                    continue  # Going back to the validation cycle

                elif values['IN3'] == '' or values['IN3'] is None:  # Making sure that the user has entered the total budget and hasn't closed the window
                    window2 = sg.popup("You didn't enter the budget.", title='', background_color='blue', icon=ICON)  # Displaying the relevant message
                    key = 0  # Updating the value of 'key'
                    event, values = window.read()  # Receiving input again
                    if values != None:
                        update(window['IN2'], values['IN2']) 
                    continue  # Going back to the validation cycle

                if table.iloc[[0],[0, 1, 2, 3, 4, 5]].isnull().values.all():  # Checking whether the first row is empty or not
                    key5 = 0  # Updating the value of 'key5'
                    table.iloc[0, 0] = 2  # Assigning a default value to the cell in the first row and the 'after_pump' column
                    table.iloc[0, 1] = 1  # Assigning a default value to the cell in the first row and the 'pressure_relief' column
                    table.iloc[0, 2] = 2  # Assigning a default value to the cell in the first row and the 'high consumers' column
                    table.iloc[0, 3] = 1  # Assigning a default value to the cell in the first row and the 'critical points' column
                    table.iloc[0, 4] = 1  # Assigning a default value to the cell in the first row and the 'village bifurcations' column
                    table.iloc[0, 5] = 1  # Assigning a default value to the cell in the first row and the 'control valve' column

                table2 = table.isnull()  # Making a boolean dataframe stating whether different entries are missing values or not, using the '.isnull' method
                for i in range(0, length):  # We have assumed that our desired excel file can have as many rows(records) as the user wants.
                    if i not in J:  # Checking whether index 'i' is among the empty rows indices or not
                        for j in range(0, 8):  # We have assumed that our desired excel file should have 8 columns.
                            if table2.iloc[i, j]:  # Finding the 'NaN' entries(i.e missing values)
                                if i != 0 or j != 7:  # Checking whether we are not in the first row or in the eighth column, or not
                                    if i == 0 and j != 7 and j != 6 and key5 == 1:  # Checking whether we are in the first row and not in the seventh or eighth column and
                                        # the first row is empty or not
                                        key = 0  # Updating the value of 'key'
                                        key6 = 0  # Updating the value of 'key6'
                                        if event == 'OK' and table2.iloc[i, j] == 1 :  # Checking whether the cell is still empty or not
                                            window2 = sg.popup('The cell at row', i + 2, 'and column', j + 1, 'is empty.', title='', background_color='blue', icon=ICON)  # Displaying the relevant message
                                        break  # Getting out of the first loop since there is a problem with the inputs
                                    
                                    else:
                                        if i != 0 or j != 6:  # Checking whether we are not in the first row or in the seventh column, or not
                                            key = 0  # Updating the value of 'key'
                                            key6 = 0  # Updating the value of 'key6'
                                            if event == 'OK' and table2.iloc[i, j] == 1 :  # Checking whether the cell is still empty or not
                                                window2 = sg.popup('The cell at row', i + 2, 'and column', j + 1, 'is empty.', title='', background_color='blue', icon=ICON)  # Displaying the relevant message
                                            break  # Getting out of the first loop since there is a problem with the inputs
                        
                        if key6 == 0:  # Checking whether there is a problem with inputs or not
                            break  # Getting out of the second loop since there is a problem with the inputs

                if key6 == 1:  # Checking whether all the cells have been filled or not
                    for i in range(0, 7):
                        if (type(table.iloc[0, i]) != float) and (type(table.iloc[0, i]) != int) and (type(table.iloc[0, i]) != np.float64) and \
                            (type(table.iloc[0, i]) != np.int64):  # Checking whether the elements of the first row are non-floats or non-ints or not: Note the classes 'np.float64'
                            # and 'np.int64'. Pandas is built on numpy.
                            window2 = sg.popup('The entries of your first row should all be of type float. Item', table.iloc[0, i], 'is not a float.', title='', background_color='blue', icon=ICON)  # Displaying the relevant message
                            key = 0  # Updating the value of 'key'
                            event, values = window.read()  # Receiving input again
                            if values != None:
                                update(window['IN2'], values['IN2']) 
                            break  # Getting out of the loop since there is a problem with the inputs

                        elif (type(table.iloc[0, i]) == np.float64 or type(table.iloc[0, i]) == np.int64) and table.iloc[0, i] <= 0:  # Checking whether the elements of the first
                            # row are negative floats or ints, or not
                            window2 = sg.popup('The entries of your first row should be positive. Item', table.iloc[0, i], 'is not positive.', title='', background_color='blue', icon=ICON)  # Displaying the relevant message
                            key = 0  # Updating the value of 'key'
                            event, values = window.read()  # Receiving input again
                            if values != None:
                                update(window['IN2'], values['IN2'])  
                            break  # Getting out of the loop since there is a problem with the inputs
                    if key == 0:  # Checking whether there is a problem with inputs or not
                        continue  # Going back to the validation cycle

                    J2 = list(range(1, length))  # Creating a list of natural numbers from 1 to 'length' - 1
                    for i in J:
                        J2.remove(i)  # Removing the indices of empty rows
                    table3 = table.iloc[J2, :]  # Making a dataframe with only non-empty rows

                    
                    for i in table3['after_pump']:
                        if i != 0 and i != 1:  # Checking whether the entries of the 'after_pump' column are all binaries or not
                            window2 = sg.popup('The لوله ھای بعد از مخزن یا پمپ column entries of your excel file should all be of type binary. Item', i,'is not a binary.', title='', background_color='blue', icon=ICON)  # Informing the
                            # user of the inappropriate pieces of data
                            key = 0  # Updating the value of 'key'
                            event, values = window.read()  # Receiving input again
                            if values != None:
                                update(window['IN2'], values['IN2']) 
                            break  # Getting out of the loop since there is a problem with the inputs
                    if key == 0:  # Checking whether there is a problem with inputs or not
                        continue  # Going back to the validation cycle

                    for i in table3['pressure_relief']:
                        if i != 0 and i != 1:  # Checking whether the entries of the 'pressure_relief' column are all binaries or not
                            window2 = sg.popup('The لوله ھای قبل یا بعد شیر فشارشکن column entries of your excel file should all be of type binary. Item', i,'is not an binary.', title='', background_color='blue', icon=ICON)  
                            # Informing the user of the inappropriate pieces of data
                            key = 0  # Updating the value of 'key'
                            event, values = window.read()  # Receiving input again
                            if values != None:
                                update(window['IN2'], values['IN2']) 
                            break  # Getting out of the loop since there is a problem with the inputs
                    if key == 0:  # Checking whether there is a problem with inputs or not
                        continue  # Going back to the validation cycle

                    for i in table3['high consumers']:
                        if type(i) != float and type(i) != int:  # Checking whether the entries of the 'high consumers' column are all (non-int) floats or not
                            window2 = sg.popup('The مشترکین پرمصرف column entries of your excel file should all be of type float. Item', i,'is not a float.', title='', background_color='blue', icon=ICON)  # Informing the user of the 
                            # inappropriate pieces of data
                            key = 0  # Updating the value of 'key'
                            event, values = window.read()  # Receiving input again
                            if values != None:
                                update(window['IN2'], values['IN2'])  
                            break  # Getting out of the loop since there is a problem with the inputs
                        elif abs(i) > 1:  # Checking whether the absolute value of 'i' is greater than 1 or not
                            window2 = sg.popup('The absolute value of your مشترکین پرمصرف inputs should be between 0 and 1. Item', i, "'s modulus is greater than 1.", title='', background_color='blue', icon=ICON)  # Informing the user 
                            # of the inappropriate pieces of data
                            key = 0  # Updating the value of 'key'
                            event, values = window.read()  # Receiving input again
                            if values != None:
                                update(window['IN2'], values['IN2'])
                            break  # Getting out of the loop since there is a problem with the inputs
                        elif i < 0:  # Checking whether 'i' is negative or not
                            window2 = sg.popup('The absolute value of your مشترکین پرمصرف inputs should be between 0 and 1. Item', i,"'s modulus is less than 1, but", i,'is negative', title='', background_color='blue', icon=ICON)  
                            # Informing the user of the inappropriate pieces of data
                            key = 0  # Updating the value of 'key'
                            event, values = window.read()  # Receiving input again
                            if values != None:
                                update(window['IN2'], values['IN2']) 
                            break  # Getting out of the loop since there is a problem with the inputs
                    if key == 0:  # Checking whether there is a problem with inputs or not
                        continue  # Going back to the validation cycle

                    for i in table3['critical points']:
                        if i != 0 and i != 1:  # Checking whether the entries of the 'sensitive points' column are all binaries or not
                            window2 = sg.popup('The نقاط حساس شبکه column entries of your excel file should all be of type binary. Item', i, 'is not a binary.', title='', background_color='blue', icon=ICON)  # Informing the user of the 
                            # inappropriate pieces of data
                            key = 0  # Updating the value of 'key'
                            event, values = window.read()  # Receiving input again
                            if values != None:
                                update(window['IN2'], values['IN2']) 
                            break  # Getting out of the loop since there is a problem with the inputs
                    if key == 0:  # Checking whether there is a problem with inputs or not
                        continue  # Going back to the validation cycle

                    for i in table3['village bifurcations']:
                        if i != 0 and i != 1:  # Checking whether the entries of the 'village bifurcations' column are all binaries or not
                            window2 = sg.popup('The محل انشعابات روستایی column entries of your excel file should all be of type binary. Item', i,'is not a binary.', title='', background_color='blue', icon=ICON)  # Informing the user 
                            # of the inappropriate pieces of data
                            key = 0  # Updating the value of 'key'
                            event, values = window.read()  # Receiving input again
                            if values != None:
                                update(window['IN2'], values['IN2']) 
                            break  # Getting out of the loop since there is a problem with the inputs
                    if key == 0:  # Checking whether there is a problem with inputs or not
                        continue  # Going back to the validation cycle
                    
                    for i in table3['control valve']:
                        if i != 0 and i != 1:  # Checking whether the entries of the 'control valve' column are all binaries or not
                            window2 = sg.popup('The لوله ھای مجاور با شیر کنترل دبی column entries of your excel file should all be of type binary. Item', i,'is not a binary.', title='', background_color='blue', icon=ICON)  # Informing 
                            # the user of the inappropriate pieces of data
                            key = 0  # Updating the value of 'key'
                            event, values = window.read()  # Receiving input again
                            if values != None:
                                update(window['IN2'], values['IN2'])  
                            break  # Getting out of the loop since there is a problem with the inputs
                    if key == 0:    # Checking whether there is a problem with inputs or not
                        continue  # Going back to the validation cycle

                    for i in table3['installation cost']:
                        if type(i) != float and type(i) != int:  # Checking whether the entries of the 'installation cost' column are all floats or not
                            window2 = sg.popup('The هزینه نصب column entries of your excel file should all be of type float. Item', i,'is not a float.', title='', background_color='blue', icon=ICON)  # Informing the user of the 
                            # inappropriate pieces of data
                            key = 0  # Updating the value of 'key'
                            event, values = window.read()  # Receiving input again
                            if values != None:
                                update(window['IN2'], values['IN2'])
                            break  # Getting out of the loop since there is a problem with the inputs
                        elif i <= 0:  # Checking whether 'i' is less than 0 or not
                            window2 = sg.popup('The value of your هزینه نصب inputs should be positive. Item', i, 'is not positive.', title='', background_color='blue', icon=ICON)  # Informing the user of the inappropriate pieces of data
                            key = 0  # Updating the value of 'key'
                            event, values = window.read()  # Receiving input again
                            if values != None:
                                update(window['IN2'], values['IN2'])
                            break  # Getting out of the loop since there is a problem with the inputs
                    if key == 0:  # Checking whether there is a problem with inputs or not
                        continue  # Going back to the validation cycle

                    for i in table3['index']:
                        if i == None:  # Checking whether the entries of the 'index' column are all non-empty strings or not
                            window2 = sg.popup('The نمايه column entries of your file can be of any type but should be nonempty. Some items are empty', title='', background_color='blue', icon=ICON)  # Displaying the relevant message
                            key = 0  # Updating the value of 'key'
                            break  # Getting out of the loop since there is a problem with the inputs

                    if key == 1:  # Checking whether there is no problem with the file, or not
                        
                        ###########################################################################################################################################################

                        '''
                        In this part, the optimization process is implemented.
                        '''

                        sg.ProgressBar.update(window['k'], visible=True, current_count=0)
                        sg.ProgressBar.update(window['k'],current_count = 100)  # Updating the 'current_count' attribute of the bar to 100 since the
                        # optimization problem is being solved
                        J2 = [0] + J2  # Adding the index of the first row(importance coefficients row) to 'J2'
                        table4 = table.iloc[J2, :]  # Creating a dataframe named 'table4' of nonempty rows and the importance coefficients row of of 'table'
                        I = [int(i) for i in range(1, length - step)]  # Creating a list of integers, which is to be used for iteration
                        P = [0 for i in I] + [0]  # Initializing the profit parameters
                        for i in range(2, length - step + 1):
                            P[i - 1] = table4.iloc[0, 0 ] * table4.iloc[i - 1, 0] + table4.iloc[0, 1] * table4.iloc[i - 1, 1] + table4.iloc[0, 2] * table4.iloc[i - 1, 2] + \
                            table4.iloc[0, 3] * table4.iloc[i - 1, 3] + table4.iloc[0, 4] * table4.iloc[i - 1, 4] + table4.iloc[0, 5] * table4.iloc[i - 1, 5]  # Specifying the
                            # profit of placement of each candidate point by summing up the profits with respect to all individual criterions: '.iloc' method is used to access
                            # the value of different cells of the dataframe 'table4'
                        x = pl.LpVariable.dicts(name='Decision_variables', indices=I, cat='Binary')  # Defining a family of decision variables(objects) with 'LpVariable.dicts' 
                        # method and specifying the prefix(common name) of these variables with 'name' variable, the set of indices with 'indices' variable, and the type of the 
                        # decision variables with the 'cat' variable
                        problem = pl.LpProblem(name='Sensor_Placement', sense=pl.const.LpMaximize)  # Creating a (binary) linear programming problem(object) with '.LpProblem'
                        # command and specifying its name with the 'name' variable and the type of its objective function with the 'sense' variable
                        problem += pl.lpSum([P[i] * x[i] for i in I])  #  Defining the objective function with '+=' command: '.lpSum' function is used for making a summation of
                        # some expressions which are linear with respect to the decision variables
                        problem += pl.lpSum([table4.iloc[i, 6] * x[i] for i in I]) <= float(values['IN3'])  # Specifying the constraint of the linear programming problem
                        solver = pl.PULP_CBC_CMD(logPath=os.path.realpath(__file__).removesuffix('\\gui30.py') + '\\file.log', msg=0)
                        problem.solve(solver)  # Solving the problem with '.solve' method
                        LL = []
                        for i in I:
                            if pl.value(x[i]):  # Checking whether the values of the decision variables are 1 or not with '.value' method
                                LL.append(str(table4.iloc[i, 7]))
                        window2 = sg.popup(f'You should place sensors in these locations: \n{chr(10).join(LL)}', title='', background_color='blue', icon=ICON)  # Stating where to deploy the sensors
                        sg.ProgressBar.update(window['k'], visible=False, current_count = 0)  # Updating the 'current_count' attribute of the bar to 0 since the optimization problem
                        # got solved
                    event, values = window.read()  # Receiving input again
                    if values != None:
                        update(window['IN2'], values['IN2'])
                    if event == '_FILEBROWSE_':  # Considereing the state that the user clicks the 'FileBrowse' button
                        values['_FILEBROWSE_'] = None
                        event, values = window.read()  # Receiving input again
                        if values != None:
                            update(window['IN2'], values['IN2'])
                        if event == 'OK':
                            break  # Going back to the event loop
                if event == 'OK' and key == 0:
                    event, values = window.read()  # Receiving input again
                    if values != None:
                        update(window['IN2'], values['IN2'])
            else:
                if event == '_FILEBROWSE_':  # Considereing the state that the user clicks the 'FileBrowse' button
                    values['IN2'] = None
                    event, values = window.read()  # Receiving input again
                    if values != None:
                        update(window['IN2'], values['IN2'])
                     
    if event == 'OK' and values == None:  # Checking whether the user has closed the gui window or not
        break  # Getting out of the event loop

    #########################################################################################################################################################################

    '''
    In this part, the validation tests are implemented on the browsed file.
    '''

    if event is not None:  # Checking whether the user has closed the window or not: Sometimes(for example, if we don't do anything and just close the window), when we close 
        # the window 'event' becomes 'None' other than 'sg.WIN_CLOSED'. We have put this cammand just in case.
        key3 = 1  # Updating the value of 'key3'
        key6 = 1  # Updating the value of 'key6'
        window['OK'].Update(disabled=False)  # Making the 'FileBrowse' button visible since the user has (apparently) browsed an excell file
        window['OK'].set_cursor('hand2')  # Changing the cursor of the 'OK' button to a hand
        if values['IN2'] == '':  # Checking whether the user has browsed a file or not
            window2 = sg.popup("You didn't browse anything.", title='', background_color='blue', icon=ICON)  # Displaying the relevant message: Note that if if a file is already available from the last browse and the user has clicked
            # the button but hasn't browsed a file then that last file is considered.
            window['OK'].Update(disabled=True)  # Making the 'OK' button invisible since the user hasn't browsed a file
            window['OK'].set_cursor('arrow')  # Changing the cursor of the 'OK' button to a hand
            sg.ProgressBar.update(window['k'], visible=False, current_count=0)  # Making the bar invisible and updating the 'current_count' attribute to zero since
            # the user hasn't browsed a file
            key3 = 0  # Updating the value of 'key3'
        if key3 == 1:  # Checking whether the bar should be made visible or not
            #sg.ProgressBar.update(window['k'], visible=True, current_count=0)  # Making the bar visible and updating the 'current_count' attribute to zero since
            # the user has browsed a file
            pass
        if values['IN2'] != '' and event == 'OK':  # Making sure that the user has browsed a file and hasn't closed the window and hasn't both filled and browsed
            # a sheet
            key = 1  # Initializing 'key' again just in case
            table = pd.read_excel(values['IN2'], header=0)  # Reading the input browsed excel file
            table.columns = ['after_pump', 'pressure_relief', 'high consumers', 'critical points', 'village bifurcations', 'control valve', 'installation cost', 'index']  
            # Updating the header
            
            length = len(table)  # Finding the index of the last row at which there is an entry
            step = 0  # Initializing the key for understanding how many empty rows(other than the first row) there are in the excell file
            J = []  # Initializing the list containing the indices of the empty rows of the file
            for i in range(1, length):
                if table.iloc[[i], :].isnull().values.all():  # Checking whether the (i + 1)th row is empty or not
                    step += 1  # Incrementing by 1
                    J.append(i)  # Adding this index to 'J'

            if (not values['IN1'].isnumeric() and values['IN1'] != '') or values['IN1'] == '0':  # Checking whether the input is a natural number or not
                window2 = sg.popup('The number of rows is invalid.', title='', background_color='blue', icon=ICON )  # Displaying the relevant message
                key = 0  # Updating the value of 'key'
                event, values = window.read()  # Receiving input again
                if values != None:
                    update(window['IN2'], values['IN2']) 
                continue  # Going back to the validation cycle

            elif values['IN1'] != '':  # Making sure that the user has entered the number of rows
                if (length - step - 1) != int(values['IN1']):  # Checking whether the user has counted the number of rows correctly or not
                    window2 = sg.popup('You made a mistake in counting the number of rows.', title='', background_color='blue', icon=ICON)  # Displaying the relevant message
                    window2 = sg.popup('Its number is', length - step - 1, '.', title='', background_color='blue', icon=ICON)
                    key = 0  # Updating the value of 'key'
                    event, values = window.read()  # Receiving input again
                    if values != None:
                        update(window['IN2'], values['IN2']) 
                    continue  # Going back to the validation cycle

            else:
                window2 = sg.popup("You didn't enter the number of rows.", title='', background_color='blue', icon=ICON)  # Displaying the relevant message
                key = 0  # Updating the value of 'key'
                event, values = window.read()  # Receiving input again
                if values != None:
                    update(window['IN2'], values['IN2']) 
                continue  # Going back to the validation cycle

            if (not isfloat(values['IN3']) and values['IN3'] != '') or (isfloat(values['IN3']) and float(values['IN3']) < 0):  # Checking whether the input is a nonnegative
                # real number or not
                window2 = sg.popup('Your budget is invalid.', title='', background_color='blue', icon=ICON)  # Displaying the relevant message
                key = 0  # Updating the value of 'key'
                event, values = window.read()  # Receiving input again
                if values != None:
                    update(window['IN2'], values['IN2']) 
                continue  # Going back to the validation cycle

            elif values['IN3'] == '' or values['IN3'] is None:  # Making sure that the user has entered the total budget and hasn't closed the window
                window2 = sg.popup("You didn't enter the budget.", title='', background_color='blue', icon=ICON)  # Displaying the relevant message
                key = 0  # Updating the value of 'key'
                event, values = window.read()  # Receiving input again
                if values != None:
                    update(window['IN2'], values['IN2']) 
                continue  # Going back to the validation cycle

            if table.iloc[[0],[0, 1, 2, 3, 4, 5]].isnull().values.all():  # Checking whether the first row is empty or not
                key5 = 0  # Updating the value of 'key5'
                table.iloc[0, 0] = 2  # Assigning a default value to the cell in the first row and the 'after_pump' column
                table.iloc[0, 1] = 1  # Assigning a default value to the cell in the first row and the 'pressure_relief' column
                table.iloc[0, 2] = 2  # Assigning a default value to the cell in the first row and the 'high consumers' column
                table.iloc[0, 3] = 1  # Assigning a default value to the cell in the first row and the 'critical points' column
                table.iloc[0, 4] = 1  # Assigning a default value to the cell in the first row and the 'village bifurcations' column
                table.iloc[0, 5] = 1  # Assigning a default value to the cell in the first row and the 'control valve' column

            table2 = table.isnull()  # Making a boolean dataframe stating whether different entries are missing values or not
            for i in range(0, length):  # We have assumed that our desired excel file can have as many rows as the user wants.
                if i not in J:  # Checking whether index 'i' is among the empty rows indices or not
                    for j in range(0, 8):  # We have assumed that our desired excel file should have 8 columns.
                        if table2.iloc[i, j]:  # Finding the 'NaN' entries
                            if i != 0 or j != 7:  # Checking whether we are not in the first row or in the eighth column, or not
                                if i == 0 and j != 7 and j != 6 and key5 == 1:  # Checking whether we are in the first row and not in the seventh or eighth column and the
                                    # first row is empty or not
                                    key = 0  # Updating the value of 'key'
                                    key6 = 0  # Updating the value of 'key6'
                                    if event == 'OK' and table2.iloc[i, j] == 1 :  # Checking whether the cell is still empty or not
                                        window2 = sg.popup('The cell at row', i + 2, 'and column', j + 1, 'is empty.', title='', background_color='blue', icon=ICON)  # Displaying the relevant message
                                    break  # Getting out of the first loop since there is a problem with the inputs

                                else:
                                    if i != 0 or j != 6:  # Checking whether we are not in the first row or in the seventh column, or not
                                        key = 0  # Updating the value of 'key'
                                        key6 = 0  # Updating the value of 'key'
                                        if event == 'OK' and table2.iloc[i, j] == 1 :  # Checking whether the cell is still empty or not
                                            window2 = sg.popup('The cell at row', i + 2, 'and column', j + 1, 'is empty.', title='', background_color='blue', icon=ICON)  # Displaying the relevant message
                                        break # Getting out of the first loop since there is a problem with the inputs
                    
                    if key6 == 0:  # Checking whether there is a problem with inputs or not
                        break  # Getting out of the second loop since there is a problem with the inputs

            
            if key6 == 1:  # Checking whether all the cells have been filled or not
                for i in range(0, 7):
                    if (type(table.iloc[0, i]) != float) and (type(table.iloc[0, i]) != int) and (type(table.iloc[0, i]) != np.float64) and (type(table.iloc[0, i]) != np.int64):
                        # Checking whether the elements of the first row are non-floats or non-ints or not
                        window2 = sg.popup('The entries of your first row should all be of type float. Item', table.iloc[0, i], 'is not a float.', title='', background_color='blue', icon=ICON)  # Displaying the relevant message
                        key = 0  # Updating the value of 'key'
                        event, values = window.read()  # Receiving input again
                        if values != None:
                            update(window['IN2'], values['IN2'])
                        break  # Getting out of the loop since there is a problem with the inputs
                    
                    elif (type(table.iloc[0, i]) == float or type(table.iloc[0, i]) == int or type(table.iloc[0, i]) == np.float64 or type(table.iloc[0, i]) == np.int64 ) and \
                        table.iloc[0, i] <= 0:  # Checking whether the elements of the first row are negative floats or ints, or not
                        window2 = sg.popup('The entries of your first row should be positive. Item', table.iloc[0, i], 'is not positive.', title='', background_color='blue', icon=ICON)  # Displaying the relevant message
                        key = 0  # Updating the value of 'key'
                        event, values = window.read()  # Receiving input again
                        if values != None:
                            update(window['IN2'], values['IN2']) 
                        break  # Getting out of the loop since there is a problem with the inputs
                if key == 0:  # Checking whether there is a problem with inputs or not
                    continue  # Going back to the validation cycle

                J2 = list(range(1, length))  # Creating a list of natural numbers from 1 to 'length' - 1
                for i in J:
                    J2.remove(i)  # Removing the indices of empty rows
                table3 = table.iloc[J2, :]  # Making a dataframe with only non-empty rows

                for i in table3['after_pump']:
                    if i != 0 and i != 1:  # Checking whether the entries of the 'after_pump' column are all binaries or not
                        window2 = sg.popup('The لوله ھای بعد از مخزن یا پمپ column entries of your excel file should all be of type binary. Item', i,'is not a binary.', title='', background_color='blue', icon=ICON)  # Informing the
                        # user of the inappropriate pieces of data
                        key = 0  # Updating the value of 'key'
                        event, values = window.read()  # Receiving input again
                        if values != None:
                            update(window['IN2'], values['IN2'])
                        break  # Getting out of the loop since there is a problem with the inputs
                if key == 0:  # Checking whether there is a problem with inputs or not
                    continue  # Going back to the validation cycle

                for i in table3['pressure_relief']:
                    if i != 0 and i != 1:  # Checking whether the entries of the 'pressure_relief' column are all binaries or not
                        window2 = sg.popup('The لوله ھای قبل یا بعد شیر فشارشکن column entries of your excel file should all be of type binary. Item', i,'is not an binary.', title='', background_color='blue', icon=ICON)
                        # Informing the user of the inappropriate pieces of data
                        key = 0  # Updating the value of 'key'
                        event, values = window.read()  # Receiving input again
                        if values != None:
                            update(window['IN2'], values['IN2']) 
                        break  # Getting out of the loop since there is a problem with the inputs
                if key == 0:  # Checking whether there is a problem with inputs or not
                    continue  # Going back to the validation cycle

                for i in table3['high consumers']:
                    if type(i) != float and type(i) != int:  # Checking whether the entries of the 'high consumers' column are all (non-int) floats or not
                        window2 = sg.popup('The مشترکین پرمصرف column entries of your excel file should all be of type float. Item', i,'is not a float.', title='', background_color='blue', icon=ICON)  # Informing the user of the
                        # inappropriate pieces of data
                        key = 0  # Updating the value of 'key'
                        event, values = window.read()  # Receiving input again
                        if values != None:
                            update(window['IN2'], values['IN2'])
                        break  # Getting out of the loop since there is a problem with the inputs
                    elif abs(i) > 1:  # Checking whether the absolute value of 'i' is greater than 1 or not
                        window2 = sg.popup('The absolute value of your مشترکین پرمصرف inputs should be between 0 and 1. Item', i, "'s modulus is greater than 1.", title='', background_color='blue', icon=ICON)  # Informing the user
                        # of the inappropriate pieces of data
                        key = 0  # Updating the value of 'key'
                        event, values = window.read()  # Receiving input again
                        if values != None:
                            update(window['IN2'], values['IN2'])
                        break  # Getting out of the loop since there is a problem with the inputs
                    elif i < 0:  # Checking whether 'i' is negative or not
                        window2 = sg.popup('The absolute value of your مشترکین پرمصرف inputs should be between 0 and 1. Item', i,"'s modulus is less than 1, but", i,'is negative.', title='', background_color='blue', icon=ICON)  
                        # Informing the user of the inappropriate pieces of data
                        key = 0  # Updating the value of 'key'
                        event, values = window.read()  # Receiving input again
                        if values != None:
                            update(window['IN2'], values['IN2'])
                        break  # Getting out of the loop since there is a problem with the inputs
                if key == 0:  # Checking whether there is a problem with inputs or not
                    continue  # Going back to the validation cycle

                for i in table3['critical points']:
                    if  i != 0 and i != 1:  # Checking whether the entries of the 'sensitive points' column are all binaries or not
                        window2 = sg.popup('The نقاط حساس شبکه column entries of your excel file should all be of type binary. Item', i, 'is not a binary.', title='', background_color='blue', icon=ICON)  # Informing the user of the
                        # inappropriate pieces of data
                        key = 0  # Updating the value of 'key'
                        event, values = window.read()  # Receiving input again
                        if values != None:
                            update(window['IN2'], values['IN2']) 
                        break  # Getting out of the loop since there is a problem with the inputs
                if key == 0:  # Checking whether there is a problem with inputs or not
                    continue  # Going back to the validation cycle

                for i in table3['village bifurcations']:
                    if i != 0 and i != 1:  # Checking whether the entries of the 'village bifurcations' column are all binaries or not
                        window2 = sg.popup('The محل انشعابات روستایی column entries of your excel file should all be of type binary. Item', i,'is not a binary.', title='', background_color='blue', icon=ICON)  # Informing the user
                        # of the inappropriate pieces of data
                        key = 0  # Updating the value of 'key'
                        event, values = window.read()  # Receiving input again
                        if values != None:
                            update(window['IN2'], values['IN2']) 
                        break  # Getting out of the loop since there is a problem with the inputs
                if key == 0:  # Checking whether there is a problem with inputs or not
                    continue  # Going back to the validation cycle

                for i in table3['control valve']:
                    if i != 0 and i != 1:  # Checking whether the entries of the 'control valve' column are all binaries or not
                        window2 = sg.popup('The لوله ھای مجاور با شیر کنترل دبی column entries of your excel file should all be of type binary. Item', i,'is not a binary.', title='', background_color='blue', icon=ICON)  # Informing
                        # the user of the inappropriate pieces of data
                        key = 0  # Updating the value of 'key'
                        event, values = window.read()  # Receiving input again
                        if values != None:
                            update(window['IN2'], values['IN2']) 
                        break  # Getting out of the loop since there is a problem with the inputs
                if key == 0:  # Checking whether there is a problem with inputs or not
                    continue  # Going back to the validation cycle

                for i in table3['installation cost']:
                    if type(i) != float and type(i) != int:  # Checking whether the entries of the 'installation cost' column are all floats or not: Note that type 'int' is not a
                        # problem.
                        window2 = sg.popup('The هزینه نصب column entries of your excel file should all be of type float. Item', i,'is not a float.', title='', background_color='blue', icon=ICON)  # Informing the user of the
                        # inappropriate pieces of data
                        key = 0  # Updating the value of 'key'
                        event, values = window.read()  # Receiving input again
                        if values != None:
                            update(window['IN2'], values['IN2'])
                        break  # Getting out of the loop since there is a problem with the inputs
                    elif i <= 0:  # Checking whether 'i' is less than 0 or not
                        window2 = sg.popup('The value of your هزینه نصب inputs should be positive. Item', i, 'is not positive.', title='', background_color='blue', icon=ICON)  # Informing the user of the inappropriate pieces of data
                        key = 0  # Updating the value of 'key'
                        event, values = window.read()  # Receiving input again
                        if values != None:
                            update(window['IN2'], values['IN2']) 
                        break  # Getting out of the loop since there is a problem with the inputs
                if key == 0:  # Checking whether there is a problem with inputs or not
                    continue  # Going back to the validation cycle

                for i in table3['index']:
                        if i == None:  # Checking whether the entries of the 'index' column are all non-empty strings or not
                            window2 = sg.popup('The نمايه column entries of your file can be of any type but should be nonempty. Some items are empty', title='', background_color='blue', icon=ICON)  # Displaying the relevant message
                            key = 0  # Updating the value of 'key'
                            break  # Getting out of the loop since there is a problem with the inputs

                if key == 1:  # Checking whether there is no problem with the file, or not

                    ###########################################################################################################################################################

                    '''
                    In this part, the optimization process is implemented.
                    '''
                    
                    sg.ProgressBar.update(window['k'], visible=True, current_count=0)
                    sg.ProgressBar.update(window['k'],current_count = 100)  # Updating the 'current_count' attribute of the bar to 100 since the
                    # optimization problem is being solved
                    J2 = [0] + J2  # Adding the index of the first row(importance coefficients row) to 'J2'
                    table4 = table.iloc[J2, :]  # Creating a dataframe named 'table4' of nonempty rows and the importance coefficients row of of 'table'
                    I = [int(i) for i in range(1, length - step)]  # Creating a list of integers, which is to be used for iteration
                    P = [0 for i in I] + [0]  # Initializing the profit parameters
                    for i in range(2, length - step + 1):
                        P[i - 1] = table4.iloc[0, 0 ] * table4.iloc[i - 1, 0] + table4.iloc[0, 1] * table4.iloc[i - 1, 1] + table4.iloc[0, 2] * table4.iloc[i - 1, 2] + \
                            table4.iloc[0, 3] * table4.iloc[i - 1, 3] + table4.iloc[0, 4] * table4.iloc[i - 1, 4] + table4.iloc[0, 5] * table4.iloc[i - 1, 5]  # Specifying the 
                            # profit of placement of each candidate point by summing up the profits with respect to all individual criterions
                    x = pl.LpVariable.dicts(name='Decision_variables', indices=I, cat='Binary')  # Defining a family of decision variables(objects) and specifying the
                    # prefix(common name) of these variables, the set of indices, and the type of the decision variables
                    problem = pl.LpProblem(name='SensorPlacement', sense=pl.const.LpMaximize)  # Creating a (binary) linear programming problem(object) and specifying its name
                    # and the type of its objective function
                    problem += pl.lpSum([P[i] * x[i] for i in I])  # Defining the objective function
                    problem += pl.lpSum([table4.iloc[i, 6] * x[i] for i in I]) <= float(values['IN3'])  # Specifying the constraint of the linear programming problem
                    #log = os.getcwd()
                    solver = pl.PULP_CBC_CMD(logPath=os.path.realpath(__file__).removesuffix('\\gui30.py') + '\\file.log', msg=0)
                    problem.solve(solver)  # Solving the problem
                    LL = []
                    for i in I:
                        if pl.value(x[i]):  # Checking whether the values of the decision variables are 1 or not
                            LL.append(str(table4.iloc[i, 7]))
                    window2 = sg.popup(f'You sould place sensors in these locations \n{chr(10).join(LL)}', title='', background_color='blue', icon=ICON)  # Stating where to deploy the sensor
                    sg.ProgressBar.update(window['k'], visible=False, current_count = 0)  # Updating the 'current_count' attribute of the bar to 0 since the optimization problem
                    # got solved
    else:
        window.close()  # Closing the window with '.close' method
        
    ###########################################################################################################################################################################
    
    '''
    In this section, the 'event loop' is made. 'Event loop' is a set of conditions whose satisfaction closes the GUI window.
    '''
    
    if event == sg.WIN_CLOSED:  # Checking whether the user has pressed the closing indicator or not
        break  # Getting out of the event loop
    if event == 'click':  # Checking whether the user has clicked the 'Click' button or not
        pass  # Do nothing and go back to the click loop
    else:
        if values['IN2'] == '':
            update(window['IN2'], 'Browse your file:')
        event, values = window.read()  # Receiving input again
        if values != None:
            update(window['IN2'], values['IN2'])

window.close()  # Closing the window

###########################################################################################################################
